import logging

from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from unidecode import unidecode

from src.config import default_settings

COLUMN_NR_TAGS = 2
COLUMN_NR_ALARM_TEXT = 3
DEFAULT_ALARM_SHEET_NAME = "DiscreteAlarms"
MAX_ROW = 500

logger = logging.getLogger(__name__)
logger.setLevel(default_settings.LOGGING_LVL_GLOBAL)


def get_alarms_from_excel(file_path: str) -> list[str]:
    """ "
    Get list of alarm tags from excel
    """
    wb = load_workbook(file_path)
    ws = wb[DEFAULT_ALARM_SHEET_NAME]
    alarm_tag_list = []
    for alarm_tag_tupple in ws.iter_rows(
        min_row=2,  # 1st row is title
        max_row=MAX_ROW,
        min_col=COLUMN_NR_TAGS,
        max_col=COLUMN_NR_TAGS,
        values_only=True,
    ):
        if alarm_tag_tupple[0] is None:
            break
        alarm_tag_list.append(alarm_tag_tupple[0])
    return alarm_tag_list


def write_alarm_texts_to_excel_file(file_path: str, alarm_texts: list) -> None:
    """
    Write list of alarm texts to excel
    """
    wb = load_workbook(file_path)
    ws = wb[DEFAULT_ALARM_SHEET_NAME]
    for i, value in enumerate(alarm_texts, start=2):  # 1st row is title
        ws.cell(row=i, column=COLUMN_NR_ALARM_TEXT).value = value
    try:
        wb.save(file_path)
    except PermissionError:
        logger.error("Unable to save file")


def create_email_alarm_text_file(file_path_excel_alarms: str, file_path_text_file: str):
    """
    Create text that is to be coppied in tia portal for sending of alarms. Save in text file.
    """
    wb = load_workbook(file_path_excel_alarms)
    ws = wb[DEFAULT_ALARM_SHEET_NAME]
    # Stop generating alarm texts after x number of empty alarms
    empty_cntr = 0
    alarm_cntr = 0
    with open(file_path_text_file, "w") as alarm_fc_file:
        for i, row_tupple in enumerate(
            ws.iter_rows(
                min_row=2,
                max_row=MAX_ROW,
                min_col=COLUMN_NR_ALARM_TEXT,
                max_col=COLUMN_NR_ALARM_TEXT,
                values_only=True,
            ),
            start=0,
        ):
            if row_tupple[0] is not None and row_tupple[0] != "<No value>":
                # remove ascents
                alarm_no_accents = unidecode(str(row_tupple[0]))
                alarm_fc_file.write(f"{i}:\r")
                alarm_fc_file.write(f"\t #\"String out\" := '{alarm_no_accents}';\r")
                alarm_cntr += 1
            else:
                empty_cntr += 1
                if empty_cntr > 5:
                    logger.info(f"No more alarm texts. Found {alarm_cntr} alarms")
                    break
        # TODO, append default and test values as well


def remove_invalid_symbols_from_alarm_names(file_path: str) -> None:
    wb = load_workbook(file_path)
    ws = wb[DEFAULT_ALARM_SHEET_NAME]
    # Loop through rows 2 to MAX_ROW in column A
    for row in range(2, MAX_ROW):
        cell = ws.cell(row=row, column=COLUMN_NR_TAGS)
        if isinstance(cell, MergedCell):
            continue  # Skip merged cells
        if cell.value and isinstance(
            cell.value, str
        ):  # Ensure the cell contains a string
            cleaned_value = (
                cell.value.replace('"', "").replace(".", "_").replace(" ", "_")
            )
            cell.value = cleaned_value
    try:
        wb.save(file_path)
    except PermissionError:
        logger.error("Unable to save file")
