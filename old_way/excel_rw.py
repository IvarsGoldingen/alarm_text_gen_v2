from openpyxl import load_workbook
from unidecode import unidecode
import logging

class ExcelRW:
    # ROW_NR_TAGS = 2
    ROW_NR_TAGS = 2
    ROW_NR_ALARM_TEXT = 3
    ALARM_SHEET_NAME = "DiscreteAlarms"
    #MAX_ROW = 500
    MAX_ROW = 500

    def __int__(self):
        pass

    @staticmethod
    def get_alarm_tags_as_list(file_path: str) -> list:
        wb = load_workbook(file_path)
        ws = wb[ExcelRW.ALARM_SHEET_NAME]
        alarm_tag_list = []
        for alarm_tag_tupple in ws.iter_rows(min_row=2,
                                             max_row=ExcelRW.MAX_ROW,
                                             min_col=ExcelRW.ROW_NR_TAGS,
                                             max_col=ExcelRW.ROW_NR_TAGS,
                                             values_only=True):
            if alarm_tag_tupple[0] is None:
                break
            alarm_tag_list.append(alarm_tag_tupple[0])
        return alarm_tag_list

    @staticmethod
    def remove_invalid_symbols_from_alarm_names(file_path: str) -> None:
        wb = load_workbook(file_path)
        ws = wb[ExcelRW.ALARM_SHEET_NAME]
        # Loop through rows 2 to 1000 in column A
        for row in range(2, ExcelRW.MAX_ROW):
            cell = ws.cell(row=row, column=ExcelRW.ROW_NR_TAGS)
            if cell.value and isinstance(cell.value, str):  # Ensure the cell contains a string
                cleaned_value = cell.value.replace('"', '').replace('.', '_')
                cell.value = cleaned_value
        wb.save(file_path)

    @staticmethod
    def write_alarm_texts_to_excel_file(file_path: str, alarm_texts: list):
        wb = load_workbook(file_path)
        ws = wb[ExcelRW.ALARM_SHEET_NAME]
        alarm_tag_list = []
        for i, value in enumerate(alarm_texts, start=2):
            ws.cell(row=i, column=ExcelRW.ROW_NR_ALARM_TEXT).value = value
        wb.save(file_path)

    @staticmethod
    def create_email_alarm_text_file(file_path_excel_alarms: str,
                                     file_path_text_file: str,
                                     print_to_console: bool = False):
        wb = load_workbook(file_path_excel_alarms)
        ws = wb[ExcelRW.ALARM_SHEET_NAME]
        empty_cntr = 0
        alarm_cntr = 0
        with open(file_path_text_file, "w") as alarm_fc_file:
            for i, row_tupple in enumerate(ws.iter_rows(min_row=2,
                                                        max_row=1000,
                                                        min_col=ExcelRW.ROW_NR_ALARM_TEXT,
                                                        max_col=ExcelRW.ROW_NR_ALARM_TEXT,
                                                        values_only=True), start=0):
                if row_tupple[0] is not None and row_tupple[0] != "<No value>":
                    # remove ascents
                    alarm_no_accents = unidecode(row_tupple[0])
                    if print_to_console:
                        print(f"{i}:")
                        print(f"\t #\"String out\" := \'{alarm_no_accents}\';")
                    alarm_fc_file.write(f"{i}:\r")
                    alarm_fc_file.write(f"\t #\"String out\" := \'{alarm_no_accents}\';\r")
                    alarm_cntr += 1
                else:
                    empty_cntr += 1
                    if empty_cntr > 5:
                        logging.info(f"No more alarm texts. Found {alarm_cntr} alarms")
                        break
